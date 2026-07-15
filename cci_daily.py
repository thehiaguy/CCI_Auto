"""CCI Daily PDF -> "SXcoal data.xlsx" automation.

Pipeline:
  1. Read the CCI Daily PDF. Article pages are text; the data tables
     (cover page + back pages) are embedded images, so they are rendered
     to PNG and transcribed by the Claude API (vision + structured output).
  2. Write the extracted values into the "Data Dump" sheet of the workbook,
     preserving its fixed layout (other sheets reference these cells by
     absolute address).
  3. For each time-series sheet, replicate the manual daily step:
     copy the bottom formula row down one row (Excel FillDown, so relative
     references shift), then freeze the old formula row's Data-Dump
     references and WORKDAY date cell to literal values.

The workbook is edited through Excel itself (xlwings/COM) so charts,
tables, comments and formats are fully preserved and formulas recalculate.

Usage:
  python cci_daily.py --pdf "C:\\path\\CCI Daily (Jul 06, 2026).pdf"
                      [--xlsx "C:\\path\\SXcoal data.xlsx"]
                      [--dry-run] [--json out.json] [--from-json in.json]
                      [--force] [--no-backup]

Requires ANTHROPIC_API_KEY in the environment (not needed with --from-json).
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import re
import shutil
import sys
from pathlib import Path

# Default workbook location: the current user's Downloads folder.
# Keep the file named exactly "SXcoal data.xlsx", or pass --xlsx with its path.
DEFAULT_XLSX = str(Path.home() / "Downloads" / "SXcoal data.xlsx")
MODEL = "claude-opus-4-8"
RENDER_SCALE = 2.2  # ~158 dpi; keeps pages under Claude's 2576px high-res limit

# Data Dump layout: (id, pdf table title, title_row, header_row, first_data, last_data, n_label_cols)
SECTIONS = [
    ("s1", "Fenwei CCI Price Index Daily", 3, 4, 5, 36, 1),
    ("s2", "Fenwei CCI Thermal Weekly Index", 38, 39, 40, 49, 1),
    ("s2b", "Fenwei CCI Thermal Index Weekly Average", 51, 52, 53, 57, 1),
    ("s3", "Fenwei CCI Coking Coal Index Daily", 59, 60, 61, 74, 1),
    ("s4", "China Spot-Futures Basis", 76, 77, 78, 80, 1),
    ("s5", "China Thermal Coal Prices", 83, 84, 85, 127, 1),
    ("s6a", "Fenwei CCI Truck Freight", 129, 130, 131, 141, 1),
    ("s6b", "China Coastal Coal Freight", 143, 144, 145, 151, 1),
    ("s6c", "Seaborne Coal Freight", 153, 154, 155, 158, 1),
    ("s7", "Chinese Coal Ports Roundup", 160, 161, 162, 207, 2),
]

# Time-series sheets driven by the Data Dump: (sheet name, gating section or None=report date)
SERIES_SHEETS = [
    ("Thermal Coal Daily", "s1"),
    ("Coking Coal Daily", "s1"),
    ("Met Coke Daily", "s1"),
    ("Thermal Coal Weekly", "s2"),
    ("Spot Futures", "s4"),
    ("Thermal Coal Price (Port Mine)", "s5"),
    ("Truck Freight", "s6a"),
    ("Coastal Coal Freight", "s6b"),
    ("Seaborne Coal Freight", "s6c"),
    ("Port Stockpile", None),
]

# Trailing AVERAGEIFS block on Thermal Coal Weekly ("Fenwei CCI Thermal Index
# Weekly Average", cols AR:BK). Its values self-compute from the daily data,
# but each row's date must follow the PDF's own weekly-average table date
# (section s2b — often a Sunday, which a WORKDAY guess can never produce).
WEEKLY_AVG_BLOCK = ("Thermal Coal Weekly", "AR", "BK", "s2b")

# Coastal Coal Freight extra that doesn't come from the PDF's data tables:
# - column W "Composite Index": fetched from the Shanghai Shipping Exchange
# The weekly block at cols AA.. (QHD-Guangzhou 60,000-70,000 DWT rate from the
# Monday freight article) is maintained MANUALLY — the code that used to
# auto-fill column AC was removed 2026-07-14 by choice, not by accident.
COASTAL_SHEET = "Coastal Coal Freight"
CBCFI_URL = "https://en.sse.net.cn/currentIndex?indexName=cbcfi"

# News-summary sections of Data Dump, refreshed from the issue's article text.
# n9/n10/n11 are rewritten in place: (title_row, header_row, first, last, n_cols).
# n8 is a multi-date comparison: a new "M/D Cap | M/D No." column pair is
# appended per survey date instead (rows fixed by region).
NEWS_SECTIONS = {
    "n9": (225, 226, 227, 241, 4),
    "n10": (243, 244, 245, 248, 5),
    "n11": (251, 252, 253, 274, 3),
}
N8_LAYOUT = (209, 210, 211, 219)  # title_row, header_row, first, last

MONTH_DATE_RE = re.compile(
    r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|"
    r"Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+\d{1,2},\s*\d{4}"
)
# Strings Excel would silently convert to numbers/dates; force them to stay text.
TEXTY_RE = re.compile(r"^[+\-]?[\d,.]+%?$|^\d{4}/\d{1,2}/\d{1,2}$")


def log(msg: str) -> None:
    print(msg, flush=True)


def load_env_file() -> None:
    """Load KEY=VALUE pairs from a .env file next to this script (if present)."""
    p = Path(__file__).with_name(".env")
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip().strip("'\""))


def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_num(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch.upper()) - 64
    return n


# --------------------------------------------------------------------------
# 1. Build the extraction spec from the workbook's current Data Dump layout
# --------------------------------------------------------------------------

def build_spec(xlsx: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "Data Dump" not in wb.sheetnames:
        sys.exit("ERROR: workbook has no 'Data Dump' sheet")
    ws = wb["Data Dump"]
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=220, max_col=12):
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c.value
    wb.close()

    spec = []
    for sid, title, title_row, header_row, first, last, n_label in SECTIONS:
        first_val_col = n_label + 1
        max_col = max(
            (c for (r, c) in grid if first <= r <= last and c > n_label), default=first_val_col
        )
        columns = []
        for c in range(first_val_col, max_col + 1):
            name = grid.get((header_row, c))
            columns.append({"col": c, "name": str(name) if name is not None else f"col{c}"})
        rows = []
        for r in range(first, last + 1):
            a = grid.get((r, 1))
            if a is None:
                continue
            has_data = any((r, c) in grid for c in range(first_val_col, max_col + 1))
            if not has_data:
                continue  # group header ("Thermal Coal", "QHD Port") or footnote row
            if n_label == 2:
                label = f"{a} | {grid.get((r, 2), '')}"
            else:
                label = str(a)
            rows.append({"row": r, "label": label})
        spec.append(
            {
                "id": sid,
                "pdf_table_title": title,
                "title_row": title_row,
                "n_label_cols": n_label,
                "columns": columns,
                "rows": rows,
            }
        )
    return spec


# --------------------------------------------------------------------------
# 2. Render the PDF's data-table pages (they are embedded images)
# --------------------------------------------------------------------------

def render_table_pages(pdf_path: Path) -> list[tuple[int, bytes]]:
    import io

    import pdfplumber
    import pypdfium2 as pdfium

    table_pages = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            big = [im for im in page.images if im["width"] > 400 and im["height"] > 300]
            if big:
                table_pages.append(i)
    if not table_pages:
        sys.exit("ERROR: no data-table (image) pages found in the PDF")

    doc = pdfium.PdfDocument(str(pdf_path))
    out = []
    for i in table_pages:
        img = doc[i].render(scale=RENDER_SCALE).to_pil()
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        out.append((i + 1, buf.getvalue()))
    log(f"Rendered {len(out)} table page(s): {[p for p, _ in out]}")
    return out


# --------------------------------------------------------------------------
# 2b. Extra source: the SSE website (Composite Index). The weekly coastal
# freight article is NOT parsed — column AC is keyed by hand (see COASTAL_SHEET).
# --------------------------------------------------------------------------


def article_pages_text(pdf_path: Path) -> str:
    """Concatenated text of the article pages (the pages that are not
    full-page table images)."""
    import pdfplumber

    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            big = [im for im in page.images if im["width"] > 400 and im["height"] > 300]
            if big:
                continue
            t = (page.extract_text() or "").strip()
            if t:
                parts.append(f"--- PDF page {i + 1} ---\n{t}")
    return "\n\n".join(parts)


def fetch_cbcfi() -> list[tuple[dt.date, float]]:
    """Fetch the CBCFI Composite Index (current + previous day) from the SSE."""
    import urllib.request

    req = urllib.request.Request(
        CBCFI_URL,
        headers={"User-Agent": "Mozilla/5.0", "Referer": "https://en.sse.net.cn/indices/cbcfinew.jsp"},
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        payload = json.loads(r.read().decode("utf-8"))
    d = payload["data"]
    comp = next(
        line for line in d["lineDataList"]
        if line.get("dataItemTypeName") == "CBCFI_T"
        or line.get("properties", {}).get("lineName_EN") == "COMPOSITE INDEX"
    )
    out = []
    for dkey, vkey in (("lastDate", "lastContent"), ("currentDate", "currentContent")):
        ds, v = d.get(dkey), comp.get(vkey)
        if ds and v is not None:
            out.append((dt.date.fromisoformat(ds), round(float(v), 2)))
    return out


# --------------------------------------------------------------------------
# 3. Extraction via the Claude API (vision + structured output)
# --------------------------------------------------------------------------

EXTRACTION_SCHEMA = {
    "type": "object",
    "properties": {
        "report_date": {"type": "string", "description": "Issue date, YYYY-MM-DD"},
        "issue": {"type": "string", "description": "Issue number, digits only"},
        "sections": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "id": {"type": "string"},
                    "date": {"type": "string", "description": "Table date, YYYY-MM-DD"},
                    "rows": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "label": {"type": "string"},
                                "values": {
                                    "type": "array",
                                    "items": {
                                        "anyOf": [
                                            {"type": "number"},
                                            {"type": "string"},
                                            {"type": "null"},
                                        ]
                                    },
                                },
                            },
                            "required": ["label", "values"],
                            "additionalProperties": False,
                        },
                    },
                },
                "required": ["id", "date", "rows"],
                "additionalProperties": False,
            },
        },
        "news": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "id": {"type": "string"},
                    "date": {"type": "string"},
                    "period": {"anyOf": [{"type": "string"}, {"type": "null"}]},
                    "rows": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "cells": {
                                    "type": "array",
                                    "items": {
                                        "anyOf": [
                                            {"type": "string"},
                                            {"type": "number"},
                                            {"type": "null"},
                                        ]
                                    },
                                }
                            },
                            "required": ["cells"],
                            "additionalProperties": False,
                        },
                    },
                },
                "required": ["id", "date", "period", "rows"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["report_date", "issue", "sections", "news"],
    "additionalProperties": False,
}

PROMPT = """These are page images from one issue of the "CCI Daily" coal-market PDF (Sxcoal/Fenwei).
Transcribe its data tables into JSON following the extraction spec below.

The spec lists, for each section: the table title as it appears in the PDF (`pdf_table_title`),
the column headers to extract (`columns`, in order), and the exact row labels expected (`rows`).

Rules:
- Return one section object per spec section, using the spec `id`.
- Return one row object per spec row label, using the spec label VERBATIM, in spec order.
  For the ports roundup the label is "Port | Metric" (e.g. "Qinhuangdao | Stockpile").
- `values` must align 1:1 with the spec's `columns` list for that section.
- Plain numbers: JSON numbers. Strip thousands separators and a leading "+" but keep negative signs.
- Cells displayed with a % sign (e.g. -23.7%): return the display string, e.g. "-23.7%".
- Text cells (Basis, Vessel Size, etc.): return the text as shown.
- The Date column in the ports roundup: string "YYYY/MM/DD".
- Blank cells: null.
- `date` per section: the date in that table's own title ("... on Jul 03, 2026"); if the title
  has no date, use the report date.
- `report_date` and `issue`: from the cover page header ("CCI Daily, Issue NNNN, <date>").
- If a spec row is missing from the PDF, return it with all-null values.
- Transcribe carefully; do not round or infer values.

EXTRACTION SPEC:
"""

NEWS_PROMPT = """

ADDITIONALLY: refresh the workbook's news-summary sections from the ARTICLE TEXT
below, adding a top-level "news" array to the same JSON object. Each entry:
{"id": "...", "date": "YYYY-MM-DD", "period": "..." or null, "rows": [{"cells": [...]}]}

Sections (omit an entry entirely if the issue has no matching content):
- "n8" Shanxi coal-mine suspensions: ONLY if an article reports Shanxi mine
  suspension/stoppage figures broken down by region (suspended capacity and
  number of mines). One row per region as printed, cells =
  [Region/City, County ("" if none), suspended capacity in Mtpa (number),
  number of mines (number)]. Include a "Total" row if given.
  "date" = the survey date the figures refer to.
- "n9" Mongolian & import coking coal update: up to 15 key items about
  Mongolian / Russian / imported coking coal, port ex-stock prices, coking &
  thermal mine surveys (utilisation, stocks), mine-mouth price moves. cells =
  [Item / Location, Value / Status, WoW or DoD Change, Notes]. Include the
  date of each figure inside the item text like the examples "(Jun 11)".
- "n10" Indonesia HBA reference prices: ONLY if the issue carries the HBA
  (ESDM) table. cells = [HBA Grade, CV Basis (Kcal/kg GAR) (number),
  Price (US$/t) (number), Change (US$/t), Change (%)].
  "period" = the price period, e.g. "H1 July 2026".
- "n11" Global market notes: up to 20 items of international / seaborne /
  macro market news (country imports & exports, futures, tenders, freight,
  policy, strikes, weather/El Nino, power demand). cells =
  [Topic, Value / Status, Notes].

Keep cells concise (a spreadsheet summary, not prose). Numbers stay numbers.
"date" defaults to the report date.

ARTICLE TEXT:
"""


def spec_for_model(spec: list[dict]) -> list[dict]:
    return [
        {
            "id": s["id"],
            "pdf_table_title": s["pdf_table_title"],
            "columns": [c["name"] for c in s["columns"]],
            "rows": [r["label"] for r in s["rows"]],
        }
        for s in spec
    ]


def run_extraction(
    images: list[tuple[int, bytes]], spec: list[dict], engine: str, articles: str = ""
) -> dict:
    if engine == "api":
        return extract_with_api(images, spec, articles)
    return extract_with_cli(images, spec, articles)


def extract_with_cli(images: list[tuple[int, bytes]], spec: list[dict], articles: str = "") -> dict:
    """Extract via the local `claude` CLI — uses the Claude subscription, no API bill."""
    import shutil as _shutil
    import subprocess
    import tempfile

    claude = _shutil.which("claude")
    if not claude:
        sys.exit("ERROR: 'claude' CLI not found on PATH (install Claude Code, or use --engine api)")

    with tempfile.TemporaryDirectory(prefix="cci_") as td:
        paths = []
        for page_no, png in images:
            p = Path(td) / f"page_{page_no:02d}.png"
            p.write_bytes(png)
            paths.append(str(p))

        prompt = (
            "Read ALL of these PNG image files (each is one page of a PDF report):\n"
            + "\n".join(paths)
            + "\n\n"
            + PROMPT
            + json.dumps(spec_for_model(spec), indent=1)
            + (NEWS_PROMPT + articles if articles else "")
            + "\n\nOUTPUT FORMAT: respond with ONLY a single JSON object — no markdown fences,"
            " no commentary before or after. Shape:\n"
            '{"report_date": "YYYY-MM-DD", "issue": "NNNN", "sections":'
            ' [{"id": "...", "date": "YYYY-MM-DD", "rows":'
            ' [{"label": "...", "values": [number|string|null, ...]}]}],'
            ' "news": [{"id": "...", "date": "YYYY-MM-DD", "period": "..."|null,'
            ' "rows": [{"cells": [string|number|null, ...]}]}]}'
        )

        log("Calling the claude CLI (uses your Claude subscription; may take a few minutes) ...")
        r = subprocess.run(
            [claude, "-p", "--output-format", "json", "--allowedTools", "Read"],
            input=prompt,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=1800,
            cwd=td,
        )
    if r.returncode != 0:
        sys.exit(f"ERROR: claude CLI failed (exit {r.returncode}):\n{(r.stderr or r.stdout)[:800]}")
    try:
        envelope = json.loads(r.stdout)
        text = envelope.get("result", "") if isinstance(envelope, dict) else r.stdout
    except json.JSONDecodeError:
        text = r.stdout
    start, end = text.find("{"), text.rfind("}")
    if start == -1 or end <= start:
        sys.exit(f"ERROR: no JSON found in claude CLI output:\n{text[:800]}")
    data = json.loads(text[start : end + 1])
    log("Extraction done (claude CLI).")
    return data


def extract_with_api(images: list[tuple[int, bytes]], spec: list[dict], articles: str = "") -> dict:
    """Extract via the Anthropic API — requires ANTHROPIC_API_KEY (pay per use)."""
    import anthropic

    client = anthropic.Anthropic()

    content = []
    for page_no, png in images:
        content.append({"type": "text", "text": f"PDF page {page_no}:"})
        content.append(
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": base64.standard_b64encode(png).decode(),
                },
            }
        )
    content.append(
        {
            "type": "text",
            "text": PROMPT + json.dumps(spec_for_model(spec), indent=1)
            + (NEWS_PROMPT + articles if articles else ""),
        }
    )

    log(f"Calling Claude ({MODEL}) to transcribe the tables ...")
    with client.messages.stream(
        model=MODEL,
        max_tokens=32000,
        thinking={"type": "adaptive"},
        output_config={"format": {"type": "json_schema", "schema": EXTRACTION_SCHEMA}},
        messages=[{"role": "user", "content": content}],
    ) as stream:
        msg = stream.get_final_message()

    if msg.stop_reason == "max_tokens":
        sys.exit("ERROR: extraction output was truncated (max_tokens); rerun or raise the limit")
    text = next(b.text for b in msg.content if b.type == "text")
    data = json.loads(text)
    u = msg.usage
    log(f"Extraction done (input {u.input_tokens} tok, output {u.output_tokens} tok).")
    return data


def validate(data: dict, spec: list[dict]) -> None:
    by_id = {s["id"]: s for s in data.get("sections", [])}
    problems = []
    for s in spec:
        got = by_id.get(s["id"])
        if got is None:
            problems.append(f"section {s['id']} missing from extraction")
            continue
        want_labels = [r["label"] for r in s["rows"]]
        got_labels = {norm_label(r["label"]) for r in got["rows"]}
        for lbl in want_labels:
            if norm_label(lbl) not in got_labels:
                problems.append(f"{s['id']}: row not extracted: {lbl!r}")
        ncols = len(s["columns"])
        for r in got["rows"]:
            if len(r["values"]) != ncols:
                problems.append(
                    f"{s['id']} row {r['label']!r}: {len(r['values'])} values, expected {ncols}"
                )
    if problems:
        for p in problems:
            log(f"  WARNING: {p}")
    dt.date.fromisoformat(data["report_date"])  # raises if malformed


def norm_label(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip().lower()


# --------------------------------------------------------------------------
# 4. Write the workbook via Excel (xlwings) — preserves charts & formats
# --------------------------------------------------------------------------

def excel_safe(v):
    """Prefix Excel-parseable strings with an apostrophe so they stay text."""
    if isinstance(v, str) and TEXTY_RE.match(v.strip()):
        return "'" + v
    return v


def section_date(data: dict, sid: str) -> dt.date:
    d = section_date_or_none(data, sid)
    return d if d else dt.date.fromisoformat(data["report_date"])


def section_date_or_none(data: dict, sid: str) -> dt.date | None:
    for s in data["sections"]:
        if s["id"] == sid:
            try:
                return dt.date.fromisoformat(s["date"])
            except (ValueError, KeyError, TypeError):
                break
    return None


def retitle(dump, row: int, d: dt.date) -> None:
    cell = dump.range((row, 1))
    title = cell.value
    if isinstance(title, str) and MONTH_DATE_RE.search(title):
        cell.value = MONTH_DATE_RE.sub(f"{d:%B} {d.day}, {d.year}", title, count=1)


def write_news_sections(dump, news: list[dict], report_date: dt.date) -> None:
    """Refresh the news-summary sections (8-11) of Data Dump from the issue's
    articles: n9/n10/n11 are rewritten in place, n8 gains a date column pair."""
    for sec in news:
        sid = sec.get("id")
        rows = sec.get("rows") or []
        try:
            d = dt.date.fromisoformat(sec.get("date") or "")
        except ValueError:
            d = report_date

        if sid == "n8" and rows:
            t_row, h_row, first, last = N8_LAYOUT
            headers = dump.range((h_row, 1), (h_row, 60)).value
            tag = f"{d.month}/{d.day}"
            if any(isinstance(h, str) and h.startswith(tag + " ") for h in headers):
                log(f"  Data Dump n8: {tag} columns already present")
                continue
            edge = max(i for i, h in enumerate(headers) if h is not None) + 1
            dump.range((h_row, edge + 1)).value = f"{tag} Cap"
            dump.range((h_row, edge + 2)).value = f"{tag} No."
            by_ab, by_a = {}, {}
            for r in range(first, last + 1):
                a = norm_label(str(dump.range((r, 1)).value or ""))
                b = norm_label(str(dump.range((r, 2)).value or ""))
                by_ab[f"{a}|{b}"] = r
                by_a.setdefault(a, []).append(r)

            def n8_row(region, county) -> int | None:
                mr = norm_label(str(region))
                mc = norm_label(str(county or ""))
                if f"{mr}|{mc}" in by_ab:
                    return by_ab[f"{mr}|{mc}"]
                base = mr.split(" (")[0]  # "total (changzhi, ...)" -> "total"
                if not mc:
                    # a city given without county means its (total) row if one exists
                    for key in (f"{base} (total)", base):
                        if len(by_a.get(key, [])) == 1:
                            return by_a[key][0]
                return None

            n = 0
            for row in rows:
                c = (row.get("cells") or []) + [None] * 4
                r = n8_row(c[0], c[1])
                if r is None:
                    log(f"  WARNING: n8 region {c[0]!r}/{c[1]!r} has no Data Dump row; skipped")
                    continue
                dump.range((r, edge + 1)).value = excel_safe(c[2])
                dump.range((r, edge + 2)).value = excel_safe(c[3])
                n += 1
            retitle(dump, t_row, d)
            log(f"  Data Dump n8: added {tag} columns ({n} regions)")

        elif sid in NEWS_SECTIONS and rows:
            t_row, h_row, first, last, ncols = NEWS_SECTIONS[sid]
            capacity = last - first + 1
            if len(rows) > capacity:
                log(f"  WARNING: {sid}: {len(rows)} items, keeping first {capacity}")
                rows = rows[:capacity]
            grid = []
            for row in rows:
                c = list(row.get("cells") or [])[:ncols]
                c += [None] * (ncols - len(c))
                grid.append([excel_safe(v) for v in c])
            grid += [[None] * ncols for _ in range(capacity - len(grid))]
            dump.range((first, 1), (last, ncols)).value = grid
            if sid == "n10" and sec.get("period"):
                cell = dump.range((t_row, 1))
                if isinstance(cell.value, str):
                    cell.value = re.sub(r"H[12]\s+\w+\s+\d{4}", sec["period"], cell.value, count=1)
            else:
                retitle(dump, t_row, d)
            log(f"  Data Dump {sid}: wrote {len(rows)} news rows")


def write_workbook(
    xlsx: Path, spec: list[dict], data: dict, force: bool, backup: bool, web: bool = True
) -> None:
    import xlwings as xw

    report_date = dt.date.fromisoformat(data["report_date"])

    if backup:
        bdir = xlsx.parent / "backups"
        bdir.mkdir(exist_ok=True)
        stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        dest = bdir / f"{xlsx.stem} {stamp}{xlsx.suffix}"
        shutil.copy2(xlsx, dest)
        log(f"Backup: {dest}")

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        wb = app.books.open(str(xlsx), update_links=False)
        dump = wb.sheets["Data Dump"]

        # --- 4a. Data Dump values -------------------------------------------------
        sections_by_id = {s["id"]: s for s in data["sections"]}
        empty_sections = []  # sections that got no data (truncated-PDF check below)
        for s in spec:
            got = sections_by_id.get(s["id"])
            if not got:
                log(f"  SKIP section {s['id']}: not in extraction")
                empty_sections.append(s["id"])
                continue
            rows_by_label = {norm_label(r["label"]): r["values"] for r in got["rows"]}
            n_written = 0
            for r in s["rows"]:
                values = rows_by_label.get(norm_label(r["label"]))
                if values is None:
                    log(f"  WARNING: {s['id']}: no extracted values for {r['label']!r}")
                    continue
                if all(v is None for v in values):
                    continue  # row absent from this issue; keep existing values
                for c, v in zip(s["columns"], values):
                    dump.range((r["row"], c["col"])).value = excel_safe(v)
                n_written += 1
            # refresh the date in the section title
            sec_date = section_date(data, s["id"])
            tcell = dump.range((s["title_row"], 1))
            title = tcell.value
            if isinstance(title, str) and MONTH_DATE_RE.search(title):
                tcell.value = MONTH_DATE_RE.sub(sec_date.strftime("%B %d, %Y"), title, count=1)
            log(f"  Data Dump {s['id']}: wrote {n_written} rows")
            if n_written == 0:
                empty_sections.append(s["id"])

        # Truncated-PDF guard: in a healthy issue every Data Dump section gets data.
        # A *cluster* of empty sections is the tell-tale sign of a missing PDF page
        # (a single empty section can be a legitimately-absent table, so don't alarm
        # on one). This does not abort — the data that did land is still valid — it
        # just flags the run loudly so a corrected PDF can be re-run.
        if len(empty_sections) >= 2:
            log("  " + "!" * 60)
            log(f"  WARNING: {len(empty_sections)} Data Dump sections got NO data: "
                f"{', '.join(empty_sections)}")
            log("  This often means the PDF is missing a page. Check the source PDF;"
                " if a corrected copy is available, restore the pre-run backup and re-run.")
            log("  " + "!" * 60)

        # --- 4a-bis. News-summary sections from the issue's articles ---------------
        if data.get("news"):
            write_news_sections(dump, data["news"], report_date)

        # header cell A1: issue + date
        a1 = dump.range("A1")
        if isinstance(a1.value, str):
            t = re.sub(r"Issue\s+\d+", f"Issue {data.get('issue', '?')}", a1.value)
            t = MONTH_DATE_RE.sub(report_date.strftime("%b %d, %Y"), t)
            a1.value = t

        app.api.CalculateFull()

        # --- 4b. Append a dated value row to each time-series sheet ---------------
        for sheet_name, sid in SERIES_SHEETS:
            target = section_date(data, sid) if sid else report_date
            try:
                sht = wb.sheets[sheet_name]
            except Exception:
                log(f"  SKIP {sheet_name!r}: sheet not found")
                continue
            append_series_row(sht, target, force)

        # --- 4c. Weekly Average block (AR:BK) dated from the PDF's s2b table ------
        d2b = section_date_or_none(data, "s2b")
        if d2b:
            extend_weekly_avg_block(wb.sheets[WEEKLY_AVG_BLOCK[0]], d2b)

        # --- 4d. Lag-dated port rows (Guangzhou) -> their own date row ------------
        fix_dated_port_values(wb, section_date(data, "s7"))

        # --- 4e. Coastal Coal Freight: SSE Composite Index (column W) --------------
        # NB: the weekly AA.. block (incl. column AC, the 60,000-70,000 DWT rate from
        # the Monday freight article) is filled in by hand — no code writes it.
        if web:
            try:
                pairs = fetch_cbcfi()
            except Exception as e:
                pairs = []
                log(f"  WARNING: SSE Composite Index fetch failed ({e}); key column W manually")
            if pairs:
                fill_composite_index(wb.sheets[COASTAL_SHEET], pairs)

        app.api.CalculateFull()
        wb.save()
        wb.close()
        log(f"Saved: {xlsx}")
    finally:
        app.quit()


def find_formula_row(sht, col: int = 1) -> int | None:
    """Row of the bottom formula cell in the given column (the live template row)."""
    max_row = sht.used_range.last_cell.row
    formulas = sht.range((1, col), (max_row, col)).formula
    if isinstance(formulas, str):
        formulas = [[formulas]]
    for i in range(len(formulas) - 1, -1, -1):
        v = formulas[i][0] if isinstance(formulas[i], (list, tuple)) else formulas[i]
        if isinstance(v, str) and v.startswith("="):
            return i + 1
    return None


def append_series_row(sht, target_date: dt.date, force: bool) -> None:
    name = sht.name
    max_col = sht.used_range.last_cell.column
    if name == WEEKLY_AVG_BLOCK[0]:
        # the AR:BK weekly-average block is managed by extend_weekly_avg_block
        # (dated from s2b); the daily append must not touch it
        max_col = min(max_col, col_num(WEEKLY_AVG_BLOCK[1]) - 1)

    f_row = find_formula_row(sht)
    if f_row is None or f_row < 2:
        log(f"  SKIP {name}: no formula row found")
        return

    prev = sht.range((f_row - 1, 1)).value
    prev_date = prev.date() if isinstance(prev, dt.datetime) else prev
    if isinstance(prev_date, dt.date) and prev_date >= target_date and not force:
        log(f"  SKIP {name}: last row {prev_date} >= report {target_date} (already current)")
        return

    formulas = sht.range((f_row, 1), (f_row, max_col)).formula
    if isinstance(formulas, str):
        formulas = ((formulas,),)
    formulas = formulas[0] if isinstance(formulas[0], (list, tuple)) else formulas

    # copy the formula row down one row (FillDown adjusts relative references)
    last = col_letter(max_col)
    sht.range(f"A{f_row}:{last}{f_row + 1}").api.FillDown()

    # freeze the old formula row: Data Dump refs -> literal values, WORKDAY -> report date
    computed = sht.range((f_row, 1), (f_row, max_col)).value
    if not isinstance(computed, (list, tuple)):
        computed = [computed]
    n_frozen = 0
    for idx, f in enumerate(formulas):
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        cell = sht.range((f_row, idx + 1))
        if "WORKDAY(" in f.upper():
            cell.value = target_date
            n_frozen += 1
        elif "Data Dump'!" in f or "'Data Dump'" in f:
            cell.value = computed[idx]
            n_frozen += 1
        # all other formulas (diffs, SUMs, AVERAGEIFS) stay live: they only
        # reference in-sheet cells and remain stable once those are literals

    log(f"  {name}: appended {target_date} (row {f_row}, froze {n_frozen} cells)")


def extend_weekly_avg_block(sht, s2b_date: dt.date) -> None:
    """Extend the Weekly Average block (AR:BK) by one row, dated from the PDF's
    own weekly-average table (s2b). Runs on any issue that carries a new s2b
    date (normally Mondays), independent of the Friday weekly-index append."""
    _, c1, c2, _ = WEEKLY_AVG_BLOCK
    c1i = col_num(c1)

    max_row = sht.used_range.last_cell.row
    bottom = None
    for r in range(max_row, 0, -1):
        cell = sht.range((r, c1i))
        if cell.formula or cell.value is not None:
            bottom = r
            break
    if bottom is None:
        log(f"  SKIP {sht.name} weekly-average block: no rows found in {c1}")
        return

    f = sht.range((bottom, c1i)).formula
    if isinstance(f, str) and f.startswith("="):
        # legacy state: bottom date is a live WORKDAY guess; the frozen date above gates
        prev = sht.range((bottom - 1, c1i)).value
        prev_date = prev.date() if isinstance(prev, dt.datetime) else prev
        if isinstance(prev_date, dt.date) and s2b_date <= prev_date:
            return
        sht.range(f"{c1}{bottom}:{c2}{bottom + 1}").api.FillDown()
        sht.range((bottom, c1i)).value = s2b_date
        new_row = bottom
    else:
        v = sht.range((bottom, c1i)).value
        cur = v.date() if isinstance(v, dt.datetime) else v
        if isinstance(cur, dt.date) and s2b_date <= cur:
            return
        sht.range(f"{c1}{bottom}:{c2}{bottom + 1}").api.FillDown()
        sht.range((bottom + 1, c1i)).value = s2b_date
        new_row = bottom + 1
    log(f"  {sht.name}: weekly-average block dated {s2b_date} (row {new_row}, from s2b)")


def fix_dated_port_values(wb, target_date: dt.date) -> None:
    """Ports whose table row carries its own (older) date — e.g. Guangzhou,
    published with a lag — get their value moved from the report-date row of
    Port Stockpile to the row matching the date printed in the PDF."""
    first, last = next((s[4], s[5]) for s in SECTIONS if s[0] == "s7")
    dump = wb.sheets["Data Dump"]
    sht = wb.sheets["Port Stockpile"]

    f_row = find_formula_row(sht)
    if f_row is None:
        return
    max_col = sht.used_range.last_cell.column
    formulas = sht.range((f_row, 1), (f_row, max_col)).formula
    if isinstance(formulas, str):
        formulas = ((formulas,),)
    formulas = formulas[0] if isinstance(formulas[0], (list, tuple)) else formulas

    dates = sht.range((1, 1), (f_row - 1, 1)).value  # column A, for date lookups
    def sheet_row(d: dt.date) -> int | None:
        for i in range(len(dates) - 1, max(-1, len(dates) - 60), -1):
            v = dates[i]
            if isinstance(v, dt.datetime) and v.date() == d:
                return i + 1
        return None

    for idx, f in enumerate(formulas):
        if not isinstance(f, str):
            continue
        m = re.search(r"'Data Dump'!\$D\$(\d+)", f)
        if not m or not (first <= int(m.group(1)) <= last):
            continue
        drow = int(m.group(1))
        own = dump.range((drow, 3)).value  # the row's own Date cell
        if isinstance(own, dt.datetime):
            own_date = own.date()
        elif isinstance(own, str) and re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", own.strip()):
            y, mo, dy = map(int, own.strip().split("/"))
            own_date = dt.date(y, mo, dy)
        else:
            continue
        if own_date == target_date:
            continue
        val = dump.range((drow, 4)).value
        if val is None:
            continue
        r_own = sheet_row(own_date)
        if r_own is None:
            log(f"  WARNING: Port Stockpile has no {own_date} row for the lag-dated"
                f" value in Data Dump row {drow}; leaving it on the {target_date} row")
            continue
        col = idx + 1
        sht.range((r_own, col)).value = excel_safe(val)
        r_tgt = sheet_row(target_date)
        if r_tgt is not None and r_tgt != r_own:
            tgt = sht.range((r_tgt, col))
            if tgt.value == val:
                tgt.clear_contents()
        log(f"  Port Stockpile: {col_letter(col)} value {val} placed at its own date"
            f" {own_date} (row {r_own})")


def fill_composite_index(sht, pairs: list[tuple[dt.date, float]]) -> None:
    """Write the SSE CBCFI Composite Index into the matching date rows."""
    max_col = sht.used_range.last_cell.column
    headers = sht.range((4, 1), (4, max_col)).value
    if not isinstance(headers, (list, tuple)):
        headers = [headers]
    col = next(
        (i + 1 for i, h in enumerate(headers)
         if isinstance(h, str) and h.strip().lower() == "composite index"),
        None,
    )
    if col is None:
        log("  WARNING: 'Composite Index' header not found on Coastal Coal Freight")
        return

    # only fill literal dated rows: writing into the live formula row would get
    # copied down by the next day's FillDown and go stale
    f_row = find_formula_row(sht)
    max_row = (f_row - 1) if f_row else sht.used_range.last_cell.row
    dates = sht.range((1, 1), (max_row, 1)).value
    for d, v in pairs:
        row = None
        for i in range(len(dates) - 1, max(-1, len(dates) - 60), -1):
            if isinstance(dates[i], dt.datetime) and dates[i].date() == d:
                row = i + 1
                break
        if row is None:
            continue  # index published for a date the sheet doesn't track yet
        cell = sht.range((row, col))
        if cell.value is None:
            cell.value = v
            log(f"  {sht.name}: Composite Index {v} at {d} (row {row}, from SSE)")


# --------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="CCI Daily PDF -> SXcoal Excel")
    ap.add_argument("--pdf", help="path to the CCI Daily PDF")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help="path to SXcoal data.xlsx")
    ap.add_argument("--dry-run", action="store_true", help="extract only; don't touch the workbook")
    ap.add_argument("--json", help="also save the extracted data to this JSON file")
    ap.add_argument("--from-json", help="skip the PDF/API and load extraction from this JSON file")
    ap.add_argument("--force", action="store_true", help="append rows even if the date already exists")
    ap.add_argument("--no-backup", action="store_true", help="don't copy the workbook to backups\\ first")
    ap.add_argument("--no-web", action="store_true",
                    help="skip fetching the SSE Composite Index (column W stays manual)")
    ap.add_argument(
        "--engine",
        choices=["cli", "api"],
        default="cli",
        help="extraction engine: 'cli' = local claude CLI / subscription (default), 'api' = Anthropic API key",
    )
    args = ap.parse_args()

    load_env_file()
    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.exit(f"ERROR: workbook not found: {xlsx}")

    spec = build_spec(xlsx)
    log(f"Spec built from workbook: {sum(len(s['rows']) for s in spec)} data rows in {len(spec)} sections")

    if args.from_json:
        data = json.loads(Path(args.from_json).read_text(encoding="utf-8"))
    else:
        if not args.pdf:
            sys.exit("ERROR: --pdf is required (or use --from-json)")
        pdf = Path(args.pdf)
        if not pdf.exists():
            sys.exit(f"ERROR: PDF not found: {pdf}")
        images = render_table_pages(pdf)
        articles = article_pages_text(pdf)
        data = run_extraction(images, spec, args.engine, articles)

    validate(data, spec)
    log(f"Report date: {data['report_date']}  (Issue {data.get('issue')})")

    if args.json:
        Path(args.json).write_text(json.dumps(data, indent=1), encoding="utf-8")
        log(f"Extraction saved: {args.json}")

    if args.dry_run:
        log("Dry run: workbook not modified.")
        return

    write_workbook(xlsx, spec, data, force=args.force, backup=not args.no_backup,
                   web=not args.no_web)


if __name__ == "__main__":
    main()
