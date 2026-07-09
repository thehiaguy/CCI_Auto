"""Dress rehearsal for the daily run.

Simulates *tomorrow's* CCI Daily issue by taking a saved extraction
(check.json), shifting its dates forward one workday, and running the full
write pipeline against a throwaway copy of the real workbook. Then verifies:

  1. every time-series sheet gained exactly one new value row with the new date
  2. the frozen values in that row equal the Data Dump cells they came from
  3. a fresh formula row (with Data Dump references) sits below it
  4. all charts survived the save
  5. the Data Dump section titles were re-dated

Usage:
  python test_run.py             # uses check.json (free, no API call)
  python test_run.py --live      # re-extracts from the PDF via the API first
  python test_run.py --keep      # keep test_run.xlsx afterwards for inspection

Exit code 0 = all checks passed.
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))

from cci_daily import (  # noqa: E402
    DEFAULT_XLSX,
    SERIES_SHEETS,
    build_spec,
    load_env_file,
    log,
    section_date,
    validate,
    write_workbook,
)

DD_REF = re.compile(r"'Data Dump'!\$([A-Z]+)\$(\d+)")


def next_workday(d: dt.date) -> dt.date:
    d += dt.timedelta(days=1)
    while d.weekday() >= 5:
        d += dt.timedelta(days=1)
    return d


def col_to_num(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch) - 64
    return n


def workbook_newest_date(src: Path) -> dt.date:
    """Newest literal (non-formula) date in the workbook's Port Stockpile sheet."""
    import openpyxl

    wb_f = openpyxl.load_workbook(src, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws_f, ws_v = wb_f["Port Stockpile"], wb_v["Port Stockpile"]
    newest = None
    for r in range(ws_v.max_row, 0, -1):
        v = ws_v.cell(row=r, column=1).value
        f = ws_f.cell(row=r, column=1).value
        if isinstance(v, dt.datetime) and not (isinstance(f, str) and f.startswith("=")):
            newest = v.date()
            break
    wb_f.close()
    wb_v.close()
    return newest or dt.date(2000, 1, 1)


def shift_dates(data: dict, floor: dt.date) -> tuple[dict, dt.date, dt.date]:
    """Return a copy of the extraction with all report-day dates moved to the
    first workday after both the saved extraction and the workbook's newest row."""
    data = copy.deepcopy(data)
    old = dt.date.fromisoformat(data["report_date"])
    new = next_workday(max(old, floor))
    data["report_date"] = new.isoformat()
    old_slash = old.strftime("%Y/%m/%d")
    new_slash = new.strftime("%Y/%m/%d")
    for s in data["sections"]:
        if s.get("date") == old.isoformat():
            s["date"] = new.isoformat()
        for r in s["rows"]:
            r["values"] = [new_slash if v == old_slash else v for v in r["values"]]
    return data, old, new


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--live", action="store_true", help="re-extract from the newest PDF first")
    ap.add_argument("--keep", action="store_true", help="keep test_run.xlsx afterwards")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--engine", choices=["cli", "api"], default="cli")
    args = ap.parse_args()

    load_env_file()
    src = Path(args.xlsx)
    if not src.exists():
        sys.exit(f"ERROR: workbook not found: {src}")

    check_json = HERE / "check.json"
    if args.live or not check_json.exists():
        pdfs = sorted(HERE.glob("CCI Daily*.pdf")) + sorted(
            Path(r"C:\Users\res1\Downloads").glob("CCI Daily*.pdf")
        )
        if not pdfs:
            sys.exit("ERROR: no 'CCI Daily*.pdf' found for live extraction")
        from cci_daily import render_table_pages, run_extraction

        spec = build_spec(src)
        data = run_extraction(render_table_pages(pdfs[-1]), spec, args.engine)
        check_json.write_text(json.dumps(data, indent=1), encoding="utf-8")
    else:
        data = json.loads(check_json.read_text(encoding="utf-8"))
        log(f"Using saved extraction: {check_json.name}")

    spec = build_spec(src)
    validate(data, spec)

    data, old_date, sim_date = shift_dates(data, workbook_newest_date(src))
    log(f"Simulating a new issue: {old_date} -> {sim_date}\n")

    # synthetic weekly-article value to exercise the Coastal Coal Freight AA block
    data["coastal_weekly"] = {"date": sim_date.isoformat(), "qhd_gz_60_70": 33.3}

    test_file = HERE / "test_run.xlsx"
    shutil.copy2(src, test_file)
    write_workbook(test_file, spec, data, force=False, backup=False, web=False)

    # ------------------------------ verification ------------------------------
    log("\nVerifying ...")
    import openpyxl

    passed, failed = 0, 0

    def check(ok: bool, msg: str) -> None:
        nonlocal passed, failed
        if ok:
            passed += 1
            log(f"  PASS  {msg}")
        else:
            failed += 1
            log(f"  FAIL  {msg}")

    with zipfile.ZipFile(src) as z:
        charts_before = sum(1 for n in z.namelist() if n.startswith("xl/charts/chart"))
    with zipfile.ZipFile(test_file) as z:
        charts_after = sum(1 for n in z.namelist() if n.startswith("xl/charts/chart"))
    check(charts_before == charts_after, f"charts preserved ({charts_after}/{charts_before})")

    wb_f = openpyxl.load_workbook(test_file, data_only=False)  # formulas
    wb_v = openpyxl.load_workbook(test_file, data_only=True)   # cached values
    dump_v = wb_v["Data Dump"]

    for sheet_name, sid in SERIES_SHEETS:
        target = section_date(data, sid) if sid else sim_date
        ws_f, ws_v = wb_f[sheet_name], wb_v[sheet_name]

        # last dated row in column A that is a literal (not a WORKDAY formula)
        val_row = None
        for r in range(ws_v.max_row, 0, -1):
            v = ws_v.cell(row=r, column=1).value
            f = ws_f.cell(row=r, column=1).value
            if isinstance(v, dt.datetime) and not (isinstance(f, str) and f.startswith("=")):
                val_row = r
                break
        ok_date = val_row is not None and ws_v.cell(row=val_row, column=1).value.date() == target
        check(ok_date, f"{sheet_name}: newest row {val_row} dated {target}")
        if val_row is None:
            continue

        # the row below must be a fresh formula row referencing Data Dump
        frow = val_row + 1
        refs = []
        for c in range(1, ws_f.max_column + 1):
            f = ws_f.cell(row=frow, column=c).value
            if isinstance(f, str) and f.startswith("="):
                m = DD_REF.search(f)
                if m:
                    refs.append((c, col_to_num(m.group(1)), int(m.group(2))))
        check(bool(refs), f"{sheet_name}: formula row at {frow} ({len(refs)} Data Dump refs)")

        # lag-dated port values (e.g. Guangzhou) are relocated to their own date row
        relocated = []
        if sheet_name == "Port Stockpile":
            for c, dcol, drow in list(refs):
                if dcol != 4 or not (162 <= drow <= 207):
                    continue
                own = dump_v.cell(row=drow, column=3).value
                if isinstance(own, str) and re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", own.strip()):
                    y, mo, dy = map(int, own.strip().split("/"))
                    own = dt.datetime(y, mo, dy)
                if not isinstance(own, dt.datetime) or own.date() == target:
                    continue
                refs.remove((c, dcol, drow))
                r_own = next(
                    (r for r in range(val_row, 0, -1)
                     if isinstance(ws_v.cell(row=r, column=1).value, dt.datetime)
                     and ws_v.cell(row=r, column=1).value.date() == own.date()),
                    None,
                )
                expect = dump_v.cell(row=drow, column=4).value
                got_own = ws_v.cell(row=r_own, column=c).value if r_own else None
                got_new = ws_v.cell(row=val_row, column=c).value
                relocated.append((c, own.date(), expect, got_own, got_new))
            for c, own_d, expect, got_own, got_new in relocated:
                check(
                    got_own == expect and got_new is None,
                    f"{sheet_name}: lag-dated col {c} value {expect} at its own date"
                    f" {own_d} (new row left empty)",
                )

        # frozen values in the new row must equal the Data Dump cells they reference
        bad = []
        for c, dcol, drow in refs:
            expect = dump_v.cell(row=drow, column=dcol).value
            got = ws_v.cell(row=val_row, column=c).value
            if expect is None:
                continue
            if isinstance(expect, (int, float)) and isinstance(got, (int, float)):
                if abs(expect - got) > 1e-6:
                    bad.append((c, expect, got))
            elif str(expect).strip() != str(got).strip():
                bad.append((c, expect, got))
        check(not bad, f"{sheet_name}: frozen values match Data Dump ({len(refs) - len(bad)}/{len(refs)})")
        for c, e, g in bad[:5]:
            log(f"          col {c}: expected {e!r}, got {g!r}")

    title = wb_v["Data Dump"].cell(row=3, column=1).value or ""
    check(sim_date.strftime("%B %d, %Y") in str(title), f"Data Dump titles re-dated ({title!r})")

    # coastal weekly-article block: new row dated sim_date with the synthetic value
    ws_v = wb_v["Coastal Coal Freight"]
    aa_bottom = next(
        (r for r in range(ws_v.max_row, 4, -1) if ws_v.cell(row=r, column=27).value is not None),
        None,
    )
    aa_date = ws_v.cell(row=aa_bottom, column=27).value if aa_bottom else None
    ac_val = ws_v.cell(row=aa_bottom, column=29).value if aa_bottom else None
    check(
        isinstance(aa_date, dt.datetime) and aa_date.date() == sim_date and ac_val == 33.3,
        f"Coastal weekly block: row {aa_bottom} dated {aa_date} with AC={ac_val}",
    )

    wb_f.close()
    wb_v.close()

    log(f"\n{'ALL CHECKS PASSED' if failed == 0 else 'FAILURES: ' + str(failed)}  ({passed} passed)")
    if args.keep:
        log(f"Test workbook kept: {test_file}")
    else:
        test_file.unlink(missing_ok=True)
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
